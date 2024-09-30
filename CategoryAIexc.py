import openpyxl

# Wczytaj istniejący plik Excel
wb = openpyxl.load_workbook('paragony.xlsx')
ws = wb.active

# Stwórz słownik mapujący produkty na kategorie
kategorie = {
  'SerMozzarella125g': 'Ser',
  'SerMozzŚwiatow300g': 'Ser',
  'Mleko3,9 GoBio IL': 'Mleko',
  'KiełbasaFuet 170g': 'GrzesMieso',
  'Gru Kon Luz': 'Owoce',
  'Jab Pot 1,5 kg': 'Owoce',
  'Cukier Biały 1kg': 'Cukier',
  'RyżJaśminowy4x100g': 'Ryż',
  'Ryż Basmati 4x100g': 'Ryż',
  'OgórBurgNS670 400g': 'Ogórki',
  'Ogrekkiszo880 500g': 'Ogórki',
  'SosPizzMutti400g': 'Pizza',
  'KetchŁagRolesd465g': 'Sosy',
  'PomidorPaprycz500g': 'Warzywa',
  'MarchewLuz': 'Warzywa',
  'FILET Z KURC. KL. A': 'Mięso',
  'SKRZYDLO Z KURCZAK': 'Warzywa',
  'PODUDZIE Z KURCZAK': 'Warzywa',
  'MAŚLANKA NATUR 1L': 'Nabiał',
  'MLEKO ŚWIEZE 1L': 'Mleko',
  'KUKURYDZA ZL 340G': 'Warzywa',
  'BROKUL SZT.': 'Warzywa',
  # Dodaj więcej mapowań według potrzeb
}

# Dodaj nagłówek dla nowej kolumny
ws['L1'] = 'Kategoria'

# Funkcja do przypisywania kategorii
def przypisz_kategorie(pozycja):
  if pozycja is None or pozycja.strip() == '':
      return None  # Nie przypisuj kategorii, jeśli wartość jest pusta
  pozycja = str(pozycja)  # Konwersja na string
  for klucz in kategorie:
      if klucz.lower() in pozycja.lower():
          return kategorie[klucz]
  return 'Inne'

# Przypisz kategorie do każdej pozycji
for row in range(2, ws.max_row + 1):
  pozycja = ws[f'B{row}'].value
  kategoria = przypisz_kategorie(pozycja)
  ws[f'L{row}'] = kategoria

# Zapisz zmodyfikowany plik
wb.save('zmodyfikowany_plik.xlsx')